import openpyxl
import os
import sys

# Set console encoding to utf-8 to avoid encoding errors
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

file_path = r'd:\OneDrive\Antigravity\Lotto_v200\source\Lotto645.xlsx'

try:
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    data = []
    
    header = [cell.value for cell in ws[1]]
    round_idx = header.index('회차')
    date_idx = header.index('날짜')
    num_idx_list = [header.index('번호1'), header.index('번호2'), header.index('번호3'), header.index('번호4'), header.index('번호5'), header.index('번호6')]
    bonus_idx = header.index('보너스번호')

    round_idx += 1  # Offset for openpyxl
    date_idx += 1
    num_idx_list = [idx + 1 for idx in num_idx_list]
    bonus_idx += 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        round_val = row[round_idx - 1]
        try:
            round_val = int(round_val)
        except (ValueError, TypeError):
            continue
            
        if 1000 <= round_val <= 1100:
            date_val = str(row[date_idx-1]).split(' ')[0] if row[date_idx-1] else ''
            nums = [row[i-1] for i in num_idx_list]
            bonus = row[bonus_idx-1]
            data.append({'round': round_val, 'date': date_val, 'nums': nums, 'bonus': bonus})
    
    wb.close()

    if not data:
        print("No data found for rounds 1000-1100.")
    else:
        # Sort by round ascending
        data.sort(key=lambda x: x['round'])
        
        print(f"Found {len(data)} records for rounds 1000-1100:\n")
        print(f"{'회차':<6} {'날짜':<12} {'번호1':<5} {'번호2':<5} {'번호3':<5} {'번호4':<5} {'번호5':<5} {'번호6':<5} {'보너스':<5}")
        print("-" * 80)
        
        for item in data:
            nums = item['nums']
            row_str = f"{item['round']:<6} {item['date']:<12} {nums[0]:<5} {nums[1]:<5} {nums[2]:<5} {nums[3]:<5} {nums[4]:<5} {nums[5]:<5} {item['bonus']:<5}"
            print(row_str)

except Exception as e:
    print(f"Error reading file: {e}")
