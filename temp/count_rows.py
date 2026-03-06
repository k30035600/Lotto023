import openpyxl
import os

file_path = r'source/Lotto645.xlsx'

if os.path.exists(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        print(f"Max Row: {ws.max_row}")
        print(f"Data Rows (excluding header): {max(0, ws.max_row - 1)}")
    except Exception as e:
        print(f"Error reading file: {e}")
else:
    print(f"File not found: {file_path}")
