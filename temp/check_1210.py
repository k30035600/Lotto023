
import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('source/Lotto645.xlsx', read_only=True)
    ws = wb.active
    
    found = False
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(r, 1).value
        # print(f"Checking row {r}: {cell_val}")
        if cell_val == 1210 or str(cell_val) == '1210':
            print(f"FOUND 1210 at row {r}")
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            print(vals)
            found = True
            break
            
    if not found:
        print("NOT FOUND 1210 in Lotto645.xlsx")
        print(f"Max row: {ws.max_row}")
        # Print last row to see where it ends
        if ws.max_row > 1:
            last_row_vals = [ws.cell(ws.max_row, c).value for c in range(1, ws.max_column + 1)]
            print(f"Last row ({ws.max_row}): {last_row_vals}")

    wb.close()
except Exception as e:
    print(f"Error: {e}")
