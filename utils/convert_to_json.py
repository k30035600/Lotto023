
import json
import os
import openpyxl

def convert_xlsx_to_json():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    source_dir = os.path.join(base_dir, 'source')
    
    # 1. Lotto645.xlsx -> Lotto645.json
    lotto645_path = os.path.join(source_dir, 'Lotto645.xlsx')
    if os.path.exists(lotto645_path):
        try:
            wb = openpyxl.load_workbook(lotto645_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                header = rows[0]
                data = []
                for row in rows[1:]:
                    item = {}
                    for i, h in enumerate(header):
                        if h:
                            item[str(h)] = row[i]
                    # 날짜 객체 문자열 변환
                    if '날짜' in item and hasattr(item['날짜'], 'strftime'):
                        item['날짜'] = item['날짜'].strftime('%Y-%m-%d')
                    data.append(item)
                
                json_path = os.path.join(source_dir, 'Lotto645.json')
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"변환 완료: {json_path} ({len(data)}건)")
            wb.close()
        except Exception as e:
            print(f"Lotto645 변환 실패: {e}")

    # 2. Lotto023.xlsx -> Lotto023.json
    lotto023_path = os.path.join(source_dir, 'Lotto023.xlsx')
    if os.path.exists(lotto023_path):
        try:
            wb = openpyxl.load_workbook(lotto023_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                header = rows[0]
                data = []
                for row in rows[1:]:
                    item = {}
                    for i, h in enumerate(header):
                        if h:
                            item[str(h)] = row[i]
                    data.append(item)
                
                json_path = os.path.join(source_dir, 'Lotto023.json')
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"변환 완료: {json_path} ({len(data)}건)")
            wb.close()
        except Exception as e:
            print(f"Lotto023 변환 실패: {e}")

if __name__ == '__main__':
    convert_xlsx_to_json()
